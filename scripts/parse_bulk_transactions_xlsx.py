import json
import sys
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook


EXPECTED_HEADERS = {
    "transaction reference": "transactionReference",
    "beneficiary name": "beneficiaryName",
    "amount": "amount",
    "tag": "tag",
    "remark": "remark",
}


def normalize_header(value):
    return str(value or "").strip().lower()


def normalize_text(value):
    text = str(value or "").strip()
    return text or None


def parse_amount(value):
    try:
      return float(Decimal(str(value).strip()))
    except (InvalidOperation, ValueError):
      return None


def main():
    workbook = load_workbook(sys.argv[1], data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        print(json.dumps({"rows": [], "errors": ["The uploaded sheet is empty."]}))
        return

    header_map = {}
    for idx, cell in enumerate(rows[0]):
        normalized = normalize_header(cell)
        if normalized in EXPECTED_HEADERS:
            header_map[EXPECTED_HEADERS[normalized]] = idx

    missing = [
        label
        for label in [
            "transactionReference",
            "beneficiaryName",
            "amount",
            "tag",
            "remark",
        ]
        if label not in header_map
    ]

    if missing:
        print(
            json.dumps(
                {
                    "rows": [],
                    "errors": [
                        "Missing required columns: "
                        + ", ".join(missing)
                    ],
                }
            )
        )
        return

    parsed_rows = []
    errors = []

    for row_number, values in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in values):
            continue

        transaction_reference = normalize_text(values[header_map["transactionReference"]])
        beneficiary_name = normalize_text(values[header_map["beneficiaryName"]])
        amount = parse_amount(values[header_map["amount"]])
        tag = normalize_text(values[header_map["tag"]])
        remark = normalize_text(values[header_map["remark"]])

        if not transaction_reference:
            errors.append(f"Row {row_number}: Transaction Reference is required.")
            continue

        if not beneficiary_name:
            errors.append(f"Row {row_number}: Beneficiary Name is required.")
            continue

        if amount is None or amount <= 0:
            errors.append(f"Row {row_number}: Amount must be a positive number.")
            continue

        parsed_rows.append(
            {
                "transactionReference": transaction_reference,
                "beneficiaryName": beneficiary_name,
                "amount": amount,
                "tag": tag,
                "remark": remark,
            }
        )

    print(json.dumps({"rows": parsed_rows, "errors": errors}))


if __name__ == "__main__":
    main()
