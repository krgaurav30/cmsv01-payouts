import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


HEADERS = [
    "Transaction Reference",
    "Beneficiary Name",
    "Amount",
    "Tag",
    "Remark",
]

SAMPLE_ROW = [
    "INV-2026-000143",
    "Jain Ashish",
    101.00,
    "salary",
    "May payout batch",
]


def main():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bulk Transactions"

    sheet.append(HEADERS)
    sheet.append(SAMPLE_ROW)

    header_fill = PatternFill(fill_type="solid", start_color="DFFBF0", end_color="DFFBF0")
    header_font = Font(bold=True, color="0B172A")

    for column_index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = header_fill
        cell.font = header_font
        sheet.column_dimensions[cell.column_letter].width = max(len(header) + 6, 22)

    workbook.save(sys.argv[1])


if __name__ == "__main__":
    main()
